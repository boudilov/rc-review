"""Convert structure.xlsx → structure.json and history.xlsx → history.json."""

from __future__ import annotations

import json
import shutil
import zipfile
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
XLSX = ROOT / "structure.xlsx"
JSON = ROOT / "structure.json"
HISTORY_XLSX = ROOT / "history.xlsx"
HISTORY_JSON = ROOT / "history.json"
PHOTOS = ROOT / "photos"
PHOTO_COL = 6  # zero-based index of "СОТРУДНИК ФОТО"


def load_media(xlsx: Path) -> dict[str, bytes]:
    media: dict[str, bytes] = {}
    with zipfile.ZipFile(xlsx) as archive:
        for name in archive.namelist():
            if name.startswith("xl/media/"):
                media[name] = archive.read(name)
    return media


def normalize_media_key(path: str) -> str:
    key = path.lstrip("/")
    if key in ("",):
        return key
    if not key.startswith("xl/media/"):
        key = f"xl/media/{Path(key).name}"
    return key


def extract_row_photos(ws, media: dict[str, bytes], photos_dir: Path) -> dict[int, str]:
    """Map 1-based Excel row numbers to ./photos/... URLs."""
    if photos_dir.exists():
        shutil.rmtree(photos_dir)
    photos_dir.mkdir(parents=True)

    row_photos: dict[int, str] = {}
    for img in ws._images:
        anchor = img.anchor
        if not hasattr(anchor, "_from"):
            continue

        excel_row = anchor._from.row + 1
        col = anchor._from.col
        if col != PHOTO_COL:
            continue

        blob = None
        try:
            # openpyxl keeps per-image bytes in the object itself; this avoids
            # path collisions where multiple anchors resolve to the same media key.
            blob = img._data()
        except Exception:
            blob = None

        media_key = normalize_media_key(getattr(img, "path", "") or "")
        if not blob:
            blob = media.get(media_key)
        if blob is None:
            continue

        ext = Path(media_key).suffix or ".png"
        filename = f"row-{excel_row}{ext}"
        (photos_dir / filename).write_bytes(blob)
        row_photos[excel_row] = f"./photos/{filename}"

    return row_photos


def read_rows(ws) -> list[list]:
    rows: list[list] = []
    for row in ws.iter_rows(values_only=True):
        rows.append([cell if cell is not None else None for cell in row])
    return rows


def apply_photos(rows: list[list], row_photos: dict[int, str]) -> None:
    for excel_row, photo_url in row_photos.items():
        index = excel_row - 1
        if index < 0 or index >= len(rows):
            continue
        while len(rows[index]) <= PHOTO_COL:
            rows[index].append(None)
        if not rows[index][PHOTO_COL]:
            rows[index][PHOTO_COL] = photo_url


def main() -> None:
    media = load_media(XLSX)
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb.active

    row_photos = extract_row_photos(ws, media, PHOTOS)
    rows = read_rows(ws)
    apply_photos(rows, row_photos)

    JSON.write_text(json.dumps(rows, ensure_ascii=False), encoding="utf-8")
    print(f"Wrote {len(rows)} rows to {JSON.name}")
    print(f"Extracted {len(row_photos)} photo(s) to {PHOTOS.name}/")

    if HISTORY_XLSX.exists():
        wb_history = openpyxl.load_workbook(HISTORY_XLSX, data_only=True)
        history_rows = read_rows(wb_history.active)
        HISTORY_JSON.write_text(json.dumps(history_rows, ensure_ascii=False), encoding="utf-8")
        print(f"Wrote {len(history_rows)} rows to {HISTORY_JSON.name}")
    else:
        print(f"Skip {HISTORY_JSON.name}: {HISTORY_XLSX.name} not found")


if __name__ == "__main__":
    main()
